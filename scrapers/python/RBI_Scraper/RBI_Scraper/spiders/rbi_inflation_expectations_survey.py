# Need to schedule this on 10th of every even month
import json
import os
from datetime import datetime
import re
from rich import print
import scrapy
from RBI_Scraper.items import RBI_inflation_expectations_survey_excel
from scrapy.loader import ItemLoader
import pandas as pd
from openpyxl import load_workbook
from scrapy.exceptions import CloseSpider
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from seaborn.utils import relative_luminance


# AUTOTHROTTLE_ENABLED = True by default see if you want to change it
# Add more user agents in middleware if you want

class RbiInflationExpectationsSurveySpider(scrapy.Spider):
    name = "rbi_inflation_expectations_survey"
    allowed_domains = ["website.rbi.org.in"]
    start_urls = ["https://website.rbi.org.in/web/rbi/statistics/survey?category=24927098&categoryName=Inflation%20Expectations%20Survey%20of%20Households%20-%20Bi-monthly&delta=100"]
    CSV_FILE = "rbi_inflation_expectations_survey_last_date.csv"
    custom_settings = {
        'ITEM_PIPELINES': {
            'RBI_Scraper.pipelines.Rbi_inflation_expectations_survey_pipeline': 1,
            'RBI_Scraper.pipelines.Backup_Pipeline_inflation_expectations_survey': 2
        }
    }
    # Define file path
    xlsx_file_path = r"D:\Desktop\financial_data_pipeline\data\raw\RBI_data\inflation_expectations_survey.xlsx"

    def parse(self, response):
        raw_date = response.xpath('//div[@class="notification-date font-resized"]/text()').get().strip()
        extracted_date = datetime.strptime(raw_date, "%b %d, %Y").date()

        if os.path.exists(self.CSV_FILE):
            df = pd.read_csv(self.CSV_FILE)
            last_saved_date = datetime.strptime(df.iloc[0, 0], "%Y-%m-%d").date()  # Read and convert to date

            if extracted_date == last_saved_date:
                self.log(extracted_date)
                raise CloseSpider("The new Consumer Survey is yet to be uploaded by the RBI.")

        # If new date, update CSV and proceed with scraping
        df = pd.DataFrame([[extracted_date]], columns=["date"])
        df.to_csv(self.CSV_FILE, index=False)

        latest_url = response.xpath('//a[@class="mtm_list_item_heading"]/@href').get()

        yield scrapy.Request(url=latest_url, callback=self.parse_inflation_expectations_survey, cb_kwargs=dict(date=extracted_date))



    def parse_inflation_expectations_survey(self, response, date):
        tables = response.xpath('//table[@class="tablebg"]')
        sheet_titles = [
            'T1A Product-wise 3M', 'T1B Product-wise 1Y', 'T2 Household IE', 'T3A Coherence 3M', 'T3B Coherence 1Y'
        ]
        table_titles = [
            'Table 1(a): Product-wise Expectations of Prices for Three Months ahead', 'Table 1(b): Product-wise Expectations of Prices for One Year ahead',
            'Table 3: Household Inflation Expectations – Current Perception, Three Months and One Year Ahead Expectations',
            'Table 4: Households Expecting General Price Movements in Coherence with Movements in Price Expectations of Various Product Groups: Three Months Ahead and One Year Ahead'
        ]
        ######## Downloading the excel file with latest data

        relative_url = re.findall(r'href="([^"]+\.xlsx[^"]*)"', response.text)[0]
        absolute_url = f'https://website.rbi.org.in{relative_url}'
        loader = ItemLoader(item=RBI_inflation_expectations_survey_excel(), selector=response)
        loader.add_value('file_urls', absolute_url)
        yield loader.load_item()


        ######## PIPELINE CODE BELOW of only 1 line (its the starting)
        # wb = load_workbook(self.xlsx_file_path)

        # Loop through extracted elements and check for partial matches
        for table in tables:
            element = table.xpath('.//tbody/tr[1]/td//text()').get().strip()
            if re.search(rf'\b{re.escape("Product-wise Expectations of Prices for Three Months ahead")}\b', element,
                         re.IGNORECASE):

                print(f"Match found: {element}")

                # Code for match 1
                extracted_data = []
                rows = table.xpath('.//tbody/tr[td]')

                condition_met = False

                for i, row in enumerate(rows):
                    # Get all td elements in the row
                    tds = row.xpath('./td')
                    td_count = len(tds)

                    row_text = ''.join(row.xpath('.//text()').getall()).strip()
                    # Apply condition only if it hasn't been met yet
                    if not condition_met and (
                            "Survey period ended" in row_text or
                            re.search(r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}\b', row_text, re.IGNORECASE)
                    ):
                        condition_met = True  # Mark condition as met
                        date_selector = tds[-1:]
                        extracted_date = datetime.strptime(date_selector[0].xpath('string()').get().strip(), "%b-%y").date().strftime("%d-%m-%Y")

                    # Check if row has more than 6 <td> elements
                    if td_count > 6:
                        # Get text from the 1st <td> elements
                        first_1 = tds[:1]  # [1, 2]
                        first_1_texts =  [td.xpath('string()').get().strip() for td in first_1]


                        # Get the last two elements
                        last_two = tds[-2:]  # [7, 8]
                        last_two_texts = [td.xpath('string()').get().strip() for td in last_two ]

                        # Append data for Excel (Row number, Last 2 TD texts)
                        extracted_data.append([f'Row {i + 1}', *first_1_texts,*last_two_texts])

                print(extracted_data)

                # --- (Pipeline Code starts here) ---
                #
                # # Load the workbook and select the sheet
                # sheet1 = wb["T1A Product-wise 3M"]
                #
                # # Determine the last used column in the sheet1
                # last_used_col = sheet1.max_column  # Gets the last column number
                # next_col = last_used_col + 1  # The new column where we will write data
                # next_col_letter = get_column_letter(next_col)  # Get column letter
                #
                # # Read the value from row 4 of the last column and add 1
                # last_round_no = sheet1.cell(row=4, column=(last_used_col-1)).value
                # round_no = (int(last_round_no) if isinstance(last_round_no, (int, float)) else 0) + 1
                #
                # # Write round_no (Row 4) and extracted_date (Row 5)
                # sheet1.cell(row=4, column=next_col, value=round_no)
                # sheet1.cell(row=5, column=next_col, value=extracted_date)
                #
                # # Merge and center these cells with the column after them
                # merge_range1 = f"{next_col_letter}4:{get_column_letter(next_col + 1)}4"
                # merge_range2 = f"{next_col_letter}5:{get_column_letter(next_col + 1)}5"
                # sheet1.merge_cells(merge_range1)
                # sheet1.merge_cells(merge_range2)
                #
                #
                # # Starting row for writing data
                # start_row = 6
                #
                # # Assuming extracted_data is a list of lists or tuples with 2 elements each,
                # # for example: extracted_data = [['A1', 'B1'], ['A2', 'B2'], ...]
                # for i, row_data in enumerate(extracted_data):
                #     # Write first column of data to the next empty column (last_used_col + 1)
                #     sheet1.cell(row=start_row + i, column=last_used_col + 1, value=row_data[1])
                #     # Write second column of data right next to it
                #     sheet1.cell(row=start_row + i, column=last_used_col + 2, value=row_data[2])



            if re.search(rf'\b{re.escape("Product-wise Expectations of Prices for One Year ahead")}\b', element,
                         re.IGNORECASE):
                print(f"Match found: {element}")

                # Code for match 2

                extracted_data = []
                rows = table.xpath('.//tbody/tr[td]')

                condition_met = False

                for i, row in enumerate(rows):
                    # Get all td elements in the row
                    tds = row.xpath('./td')
                    td_count = len(tds)

                    row_text = ''.join(row.xpath('.//text()').getall()).strip()
                    # Apply condition only if it hasn't been met yet
                    if not condition_met and (
                            "Survey period ended" in row_text or
                            re.search(r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2}\b', row_text, re.IGNORECASE)
                    ):
                        condition_met = True  # Mark condition as met
                        date_selector = tds[-1:]
                        extracted_date = datetime.strptime(date_selector[0].xpath('string()').get().strip(), "%b-%y").date().strftime("%d-%m-%Y")

                    # Check if row has more than 6 <td> elements
                    if td_count > 6:
                        # Get text from the last two <td> elements
                        first_1 = tds[:1]  # [1, 2]
                        first_1_texts =  [td.xpath('string()').get().strip() for td in first_1]


                        # Get the last two elements
                        last_two = tds[-2:]  # [7, 8]
                        last_two_texts = [td.xpath('string()').get().strip() for td in last_two ]

                        # Append data for Excel (Row number, Last 2 TD texts)
                        extracted_data.append([f'Row {i + 1}', *first_1_texts,*last_two_texts])

                print(extracted_data)

            # --- (Pipeline Code starts here) ---

            # sheet2 = wb["T1B Product-wise 1Y"]
            #
            # # Determine the last used column in the sheet2
            # last_used_col = sheet2.max_column  # Gets the last column number
            # next_col = last_used_col + 1  # The new column where we will write data
            # next_col_letter = get_column_letter(next_col)  # Get column letter
            #
            # # Read the value from row 4 of the last column and add 1
            # last_round_no = sheet2.cell(row=4, column=last_used_col).value
            # round_no = (int(last_round_no) if isinstance(last_round_no, (int, float)) else 0) + 1
            #
            # # Write round_no (Row 4) and extracted_date (Row 5)
            # sheet2.cell(row=4, column=next_col, value=round_no)
            # sheet2.cell(row=5, column=next_col, value=extracted_date)
            #
            # # Merge and center these cells with the column after them
            # merge_range1 = f"{next_col_letter}4:{get_column_letter(next_col + 1)}4"
            # merge_range2 = f"{next_col_letter}5:{get_column_letter(next_col + 1)}5"
            # sheet2.merge_cells(merge_range1)
            # sheet2.merge_cells(merge_range2)
            #
            # # Starting row for writing data
            # start_row = 6
            #
            # # Assuming extracted_data is a list of lists or tuples with 2 elements each,
            # # for example: extracted_data = [['A1', 'B1'], ['A2', 'B2'], ...]
            # for i, row_data in enumerate(extracted_data):
            #     # Write first column of data to the next empty column (last_used_col + 1)
            #     sheet2.cell(row=start_row + i, column=last_used_col + 1, value=row_data[1])
            #     # Write second column of data right next to it
            #     sheet2.cell(row=start_row + i, column=last_used_col + 2, value=row_data[2])


            if re.search(
                    rf'\b{re.escape("Household Inflation Expectations – Current Perception, Three Months and One Year Ahead Expectations")}\b',
                    element, re.IGNORECASE):
                print(f"Match found: {element}")


                # Code for match 3

                extracted_data = []
                data = table.xpath('(.//tr[number(string(td[last()])) = number(string(td[last()]))])[last()]/td//text()').getall()

                # 1. Extract and store the date (the non-numeric value)
                date_value = datetime.strptime(data[0], "%b-%y").date().strftime("%d-%m-%Y")

                # 2. Remove the non-numeric value from the list
                numeric_list = data[1:]

                # 3. Split the numeric_list into two lists:
                #    - odd_list: Elements in the 1st, 3rd, 5th, ... positions (using 0-based slicing, these are even indices)
                #    - even_list: Elements in the 2nd, 4th, 6th, ... positions (using 0-based slicing, these are odd indices)
                odd_list = numeric_list[0::2]
                even_list = numeric_list[1::2]

                # 4. Append these lists into a combined list (first odd_list, then even_list)
                extracted_data.append([*odd_list])
                extracted_data.append([*even_list])

                # Print the results (or later use combined_lists to write to an Excel file)
                print("Date:", date_value)
                print(extracted_data)

                # --- (Pipeline Code starts here) ---
                # sheet3 = wb["T2 Household IE"]
                # # Use sheet3.max_row as the last filled row
                # last_filled_row = sheet3.max_row
                # start_row_new = last_filled_row + 1  # New data starts below the last filled row
                #
                # # Write round_no in Column B (column index 2) and merge it with the cell below
                # sheet3.merge_cells(start_row=start_row_new, start_column=2,
                #                   end_row=start_row_new + 1, end_column=2)
                # cell_round = sheet3.cell(row=start_row_new, column=2, value=round_no)
                # cell_round.alignment = Alignment(horizontal="center", vertical="center")
                #
                # # Write date_value in Column C (column index 3) and merge it with the cell below
                # sheet3.merge_cells(start_row=start_row_new, start_column=3,
                #                   end_row=start_row_new + 1, end_column=3)
                # cell_date = sheet3.cell(row=start_row_new, column=3, value=date_value)
                # cell_date.alignment = Alignment(horizontal="center", vertical="center")
                #
                # row1_data = extracted_date[0]  # First row of data (a list of values)
                # row2_data = extracted_date[1]  # Second row of data (a list of values)
                #
                # # Write row1 data into cells starting at Column D
                # for i, value in enumerate(row1_data):
                #     sheet3.cell(row=start_row_new, column=4 + i, value=value)
                #
                # # Write row2 data into cells starting at Column D,
                # # but format each value as "[#value]"
                # for i, value in enumerate(row2_data):
                #     formatted_value = f"[#{value}]"
                #     sheet3.cell(row=start_row_new + 1, column=4 + i, value=formatted_value)

            if re.search(
                    rf'\b{re.escape("Households Expecting General Price Movements in Coherence with Movements in Price Expectations of Various Product Groups: Three Months Ahead and One Year Ahead")}\b',
                    element, re.IGNORECASE):
                print(f"Match found: {element}")

                # Code for match 4
                extracted_three_month_data = []

                data = table.xpath('(//table[@class="tablebg"][5]//tr[count(td) < 2]/preceding-sibling::tr[number(string(td[last()])) = number(string(td[last()]))])[last()]/td//text()').getall()
                if data != None:


                    # 1. Extract and store the date (the non-numeric value)
                    date_value = datetime.strptime(data[0], "%b-%y").date().strftime("%d-%m-%Y")

                    # 2. Remove the non-numeric value from the list
                    numeric_list = data[1:]

                    extracted_three_month_data.append([*numeric_list])

                    # Print the results (or later use combined_lists to write to an Excel file)
                    print("Date:", date_value)
                    print(extracted_three_month_data)

                    # --- (Pipeline Code starts here) ---


                    # sheet = wb["T3A Coherence 3M"]
                    # # Use sheet.max_row as the last filled row
                    # last_filled_row = sheet.max_row
                    # start_row_new = last_filled_row + 1  # New data starts below the last filled row
                    #
                    # # Write round_no in Column B (column index 2)
                    # cell_round = sheet.cell(row=start_row_new, column=2, value=round_no)
                    #
                    # # Write date_value in Column C (column index 3)
                    # cell_date = sheet.cell(row=start_row_new, column=3, value=date_value)
                    #
                    # row1_data = extracted_three_month_data[0]  # First row of data (a list of values)
                    #
                    # # Write row1 data into cells starting at Column D
                    # for i, value in enumerate(row1_data):
                    #     sheet.cell(row=start_row_new, column=4 + i, value=value)


                else: self.log("Website Layout has changed for T3A Coherence 3M excel sheet's data previously found in "
                               "Table 4: Households Expecting General Price Movements in Coherence with "
                               "Movements in Price Expectations of Various Product Groups: Three "
                               "Months Ahead and One Year Ahead")


                extracted_1_year_data = []
                data = table.xpath('(.//tr[number(string(td[last()])) = number(string(td[last()]))])[last()]/td//text()').getall()

                # 1. Extract and store the date (the non-numeric value)
                date_value = datetime.strptime(data[0], "%b-%y").date().strftime(
                    "%d-%m-%Y")

                # 2. Remove the non-numeric value from the list
                numeric_list = data[1:]

                # 4. Append these lists into a combined list (first odd_list, then even_list)
                extracted_1_year_data.append([*numeric_list])

                # Print the results (or later use combined_lists to write to an Excel file)
                print("Date:", date_value)

                print(extracted_1_year_data)

                # --- (Pipeline Code starts here) ---

                # sheet = wb["T3B Coherence 1Y"]
                # # Use sheet.max_row as the last filled row
                # last_filled_row = sheet.max_row
                # start_row_new = last_filled_row + 1  # New data starts below the last filled row
                #
                # # Write round_no in Column B (column index 2)
                # cell_round = sheet.cell(row=start_row_new, column=2, value=round_no)
                #
                # # Write date_value in Column C (column index 3)
                # cell_date = sheet.cell(row=start_row_new, column=3, value=date_value)
                #
                # row1_data = extracted_1_year_data[0]  # First row of data (a list of values)
                #
                # # Write row1 data into cells starting at Column D
                # for i, value in enumerate(row1_data):
                #     sheet.cell(row=start_row_new, column=4 + i, value=value)



        # wb.save(self.xlsx_file_path)


        self.log(print("task completed"))

